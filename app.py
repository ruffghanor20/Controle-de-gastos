from __future__ import annotations

import os
from datetime import date, datetime
from functools import wraps
from io import BytesIO
from pathlib import Path

import openpyxl
from flask import (
    Flask,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook
from sqlalchemy import func
from werkzeug.security import check_password_hash, generate_password_hash


BASE_DIR = Path(__file__).resolve().parent
INSTANCE_DIR = BASE_DIR / "instance"
INSTANCE_DIR.mkdir(exist_ok=True)


def build_database_uri() -> str:
    env_uri = os.getenv("DATABASE_URL", "").strip()
    if env_uri:
        if env_uri.startswith("postgres://"):
            env_uri = env_uri.replace("postgres://", "postgresql://", 1)
        return env_uri
    return f"sqlite:///{INSTANCE_DIR / 'salao.db'}"


app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = build_database_uri()
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "salao-v2-secret-key")

db = SQLAlchemy(app)


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), nullable=False, unique=True)
    password_hash = db.Column(db.String(255), nullable=False)
    is_active = db.Column(db.Boolean, nullable=False, default=True)

    def set_password(self, raw_password: str) -> None:
        self.password_hash = generate_password_hash(raw_password)

    def check_password(self, raw_password: str) -> bool:
        return check_password_hash(self.password_hash, raw_password)


class Service(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    category = db.Column(db.String(80), nullable=False)
    default_price = db.Column(db.Float, nullable=False, default=0.0)
    commission_pct = db.Column(db.Float, nullable=False, default=0.0)
    active = db.Column(db.Boolean, nullable=False, default=True)

    attendances = db.relationship("Attendance", back_populates="service")

    def __repr__(self) -> str:
        return f"<Service {self.name}>"


class Collaborator(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(80), nullable=False)
    default_commission_pct = db.Column(db.Float, nullable=False, default=0.0)
    status = db.Column(db.String(20), nullable=False, default="Ativo")

    attendances = db.relationship("Attendance", back_populates="collaborator")

    def __repr__(self) -> str:
        return f"<Collaborator {self.name}>"


class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    attendance_date = db.Column(db.Date, nullable=False)
    client_name = db.Column(db.String(120), nullable=True)
    service_id = db.Column(db.Integer, db.ForeignKey("service.id"), nullable=False)
    collaborator_id = db.Column(db.Integer, db.ForeignKey("collaborator.id"), nullable=False)
    payment_method = db.Column(db.String(40), nullable=False, default="Pix")
    amount_charged = db.Column(db.Float, nullable=False)
    commission_pct_applied = db.Column(db.Float, nullable=False, default=0.0)
    commission_amount = db.Column(db.Float, nullable=False, default=0.0)
    salon_amount = db.Column(db.Float, nullable=False, default=0.0)
    notes = db.Column(db.Text, nullable=True)

    service = db.relationship("Service", back_populates="attendances")
    collaborator = db.relationship("Collaborator", back_populates="attendances")

    def __repr__(self) -> str:
        return f"<Attendance {self.id}>"


class Expense(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    expense_date = db.Column(db.Date, nullable=False)
    category = db.Column(db.String(80), nullable=False)
    description = db.Column(db.String(200), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    expense_type = db.Column(db.String(20), nullable=False, default="Variavel")
    paid_by = db.Column(db.String(120), nullable=True)

    def __repr__(self) -> str:
        return f"<Expense {self.description}>"


class ProductPurchase(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    purchase_date = db.Column(db.Date, nullable=False)
    product_name = db.Column(db.String(120), nullable=False)
    category = db.Column(db.String(80), nullable=False)
    quantity = db.Column(db.Float, nullable=False, default=1.0)
    unit_cost = db.Column(db.Float, nullable=False, default=0.0)
    total_cost = db.Column(db.Float, nullable=False, default=0.0)
    supplier = db.Column(db.String(120), nullable=True)

    def __repr__(self) -> str:
        return f"<ProductPurchase {self.product_name}>"


def parse_float(value: str | float | int | None, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("R$", "").replace(" ", "")
    if not text:
        return default

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return default


def parse_int(value: str | int | None, default: int) -> int:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return default


def parse_bool(value) -> bool:
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"1", "sim", "true", "ativo", "yes", "y", "x", "on"}


def parse_date(value: str | date | datetime | None, fallback: date | None = None) -> date:
    if not value:
        return fallback or date.today()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return fallback or date.today()


def normalize_period(year: int, month: int) -> tuple[int, int]:
    year = min(max(year, 2020), 2100)
    month = min(max(month, 1), 12)
    return year, month


def month_bounds(year: int, month: int) -> tuple[date, date]:
    year, month = normalize_period(year, month)
    start = date(year, month, 1)
    if month == 12:
        end = date(year + 1, 1, 1)
    else:
        end = date(year, month + 1, 1)
    return start, end


def month_reference_offset(year: int, month: int, offset: int) -> tuple[int, int]:
    base = (year * 12 + month - 1) + offset
    new_year = base // 12
    new_month = (base % 12) + 1
    return new_year, new_month


def calculate_monthly_metrics(year: int, month: int) -> dict:
    year, month = normalize_period(year, month)
    start, end = month_bounds(year, month)

    attendances = (
        Attendance.query.filter(Attendance.attendance_date >= start, Attendance.attendance_date < end)
        .order_by(Attendance.attendance_date.desc(), Attendance.id.desc())
        .all()
    )
    expenses = (
        Expense.query.filter(Expense.expense_date >= start, Expense.expense_date < end)
        .order_by(Expense.expense_date.desc(), Expense.id.desc())
        .all()
    )
    products = (
        ProductPurchase.query.filter(
            ProductPurchase.purchase_date >= start,
            ProductPurchase.purchase_date < end,
        )
        .order_by(ProductPurchase.purchase_date.desc(), ProductPurchase.id.desc())
        .all()
    )

    gross_revenue = sum(item.amount_charged for item in attendances)
    commissions_paid = sum(item.commission_amount for item in attendances)
    salon_revenue = sum(item.salon_amount for item in attendances)
    operational_expenses = sum(item.amount for item in expenses)
    product_costs = sum(item.total_cost for item in products)
    result_month = salon_revenue - operational_expenses - product_costs
    count_attendances = len(attendances)
    average_ticket = gross_revenue / count_attendances if count_attendances else 0.0

    service_rank = (
        db.session.query(
            Service.name.label("service_name"),
            func.count(Attendance.id).label("total_attendances"),
            func.coalesce(func.sum(Attendance.amount_charged), 0).label("gross_amount"),
            func.coalesce(func.sum(Attendance.salon_amount), 0).label("salon_amount"),
        )
        .join(Attendance, Attendance.service_id == Service.id)
        .filter(Attendance.attendance_date >= start, Attendance.attendance_date < end)
        .group_by(Service.id, Service.name)
        .order_by(func.sum(Attendance.amount_charged).desc(), func.count(Attendance.id).desc())
        .all()
    )

    collaborator_rank = (
        db.session.query(
            Collaborator.name.label("collaborator_name"),
            func.count(Attendance.id).label("total_attendances"),
            func.coalesce(func.sum(Attendance.amount_charged), 0).label("gross_amount"),
            func.coalesce(func.sum(Attendance.commission_amount), 0).label("commission_amount"),
            func.coalesce(func.sum(Attendance.salon_amount), 0).label("salon_amount"),
        )
        .join(Attendance, Attendance.collaborator_id == Collaborator.id)
        .filter(Attendance.attendance_date >= start, Attendance.attendance_date < end)
        .group_by(Collaborator.id, Collaborator.name)
        .order_by(func.sum(Attendance.amount_charged).desc(), func.count(Attendance.id).desc())
        .all()
    )

    return {
        "year": year,
        "month": month,
        "start": start,
        "end": end,
        "attendances": attendances,
        "expenses": expenses,
        "products": products,
        "gross_revenue": gross_revenue,
        "commissions_paid": commissions_paid,
        "salon_revenue": salon_revenue,
        "operational_expenses": operational_expenses,
        "product_costs": product_costs,
        "total_outflows": operational_expenses + product_costs,
        "result_month": result_month,
        "count_attendances": count_attendances,
        "average_ticket": average_ticket,
        "service_rank": service_rank,
        "collaborator_rank": collaborator_rank,
    }


def calculate_recent_trend(year: int, month: int, window: int = 6) -> dict:
    labels: list[str] = []
    gross_values: list[float] = []
    result_values: list[float] = []

    for offset in range(-(window - 1), 1):
        item_year, item_month = month_reference_offset(year, month, offset)
        metrics = calculate_monthly_metrics(item_year, item_month)
        labels.append(f"{MONTH_SHORT[item_month]}/{str(item_year)[-2:]}")
        gross_values.append(round(metrics["gross_revenue"], 2))
        result_values.append(round(metrics["result_month"], 2))

    return {
        "labels": labels,
        "gross_values": gross_values,
        "result_values": result_values,
    }


def get_current_user() -> User | None:
    user_id = session.get("user_id")
    if not user_id:
        return None
    return User.query.get(user_id)


def login_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if not get_current_user():
            flash("Faça login para acessar o sistema.", "error")
            return redirect(url_for("login", next=request.path))
        return view_func(*args, **kwargs)

    return wrapped


def ensure_default_admin() -> None:
    username = os.getenv("ADMIN_USERNAME", "admin").strip() or "admin"
    password = os.getenv("ADMIN_PASSWORD", "admin123").strip() or "admin123"
    existing = User.query.filter_by(username=username).first()
    if existing:
        return
    user = User(username=username, is_active=True)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()


def build_attendance_values(form_data) -> dict:
    service = Service.query.get_or_404(parse_int(form_data.get("service_id"), 0))
    collaborator = Collaborator.query.get_or_404(parse_int(form_data.get("collaborator_id"), 0))
    amount_charged = parse_float(form_data.get("amount_charged"), service.default_price)

    manual_pct = form_data.get("commission_pct_applied")
    if manual_pct and str(manual_pct).strip():
        commission_pct = parse_float(manual_pct)
    elif service.commission_pct > 0:
        commission_pct = service.commission_pct
    else:
        commission_pct = collaborator.default_commission_pct

    commission_amount = round(amount_charged * (commission_pct / 100), 2)
    salon_amount = round(amount_charged - commission_amount, 2)

    return {
        "attendance_date": parse_date(form_data.get("attendance_date")),
        "client_name": (form_data.get("client_name", "").strip() or None),
        "service_id": service.id,
        "collaborator_id": collaborator.id,
        "payment_method": (form_data.get("payment_method", "Pix").strip() or "Pix"),
        "amount_charged": amount_charged,
        "commission_pct_applied": commission_pct,
        "commission_amount": commission_amount,
        "salon_amount": salon_amount,
        "notes": (form_data.get("notes", "").strip() or None),
    }


def cell_values(sheet):
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(v is not None and str(v).strip() != "" for v in row):
            yield row


def import_workbook(workbook, reset: bool = False) -> dict:
    counts = {"services": 0, "collaborators": 0, "attendances": 0, "expenses": 0, "products": 0}

    if reset:
        Attendance.query.delete()
        Expense.query.delete()
        ProductPurchase.query.delete()
        Service.query.delete()
        Collaborator.query.delete()
        db.session.commit()

    if "tbServicos" in workbook.sheetnames:
        ws = workbook["tbServicos"]
        for row in cell_values(ws):
            _, name, category, price, commission, active = list(row[:6]) + [None] * max(0, 6 - len(row[:6]))
            if not name:
                continue
            normalized_name = str(name).strip()
            exists = Service.query.filter_by(name=normalized_name).first()
            if exists:
                continue
            db.session.add(
                Service(
                    name=normalized_name,
                    category=str(category or "Geral").strip(),
                    default_price=parse_float(price),
                    commission_pct=parse_float(commission),
                    active=parse_bool(active),
                )
            )
            counts["services"] += 1
        db.session.commit()

    if "tbColaboradores" in workbook.sheetnames:
        ws = workbook["tbColaboradores"]
        for row in cell_values(ws):
            _, name, role, commission, status = list(row[:5]) + [None] * max(0, 5 - len(row[:5]))
            if not name:
                continue
            normalized_name = str(name).strip()
            exists = Collaborator.query.filter_by(name=normalized_name).first()
            if exists:
                continue
            db.session.add(
                Collaborator(
                    name=normalized_name,
                    role=str(role or "Profissional").strip(),
                    default_commission_pct=parse_float(commission),
                    status=str(status or "Ativo").strip(),
                )
            )
            counts["collaborators"] += 1
        db.session.commit()

    if "tbAtendimentos" in workbook.sheetnames:
        ws = workbook["tbAtendimentos"]
        for row in cell_values(ws):
            row = list(row[:15]) + [None] * max(0, 15 - len(row[:15]))
            (
                _row_id,
                attendance_date,
                client_name,
                service_name,
                _category,
                collaborator_name,
                _role,
                amount_charged,
                payment_method,
                service_default_pct,
                collaborator_default_pct,
                applied_pct,
                commission_value,
                salon_value,
                notes,
            ) = row

            if not service_name or not collaborator_name:
                continue

            service = Service.query.filter_by(name=str(service_name).strip()).first()
            collaborator = Collaborator.query.filter_by(name=str(collaborator_name).strip()).first()
            if not service or not collaborator:
                continue

            amount = parse_float(amount_charged, service.default_price)
            if applied_pct is not None and str(applied_pct).strip():
                pct = parse_float(applied_pct)
            elif service_default_pct is not None and parse_float(service_default_pct) > 0:
                pct = parse_float(service_default_pct)
            else:
                pct = parse_float(collaborator_default_pct, collaborator.default_commission_pct)

            commission_amount = parse_float(commission_value, round(amount * (pct / 100), 2))
            salon_amount = parse_float(salon_value, round(amount - commission_amount, 2))

            db.session.add(
                Attendance(
                    attendance_date=parse_date(attendance_date),
                    client_name=str(client_name).strip() if client_name else None,
                    service_id=service.id,
                    collaborator_id=collaborator.id,
                    payment_method=str(payment_method or "Pix").strip(),
                    amount_charged=amount,
                    commission_pct_applied=pct,
                    commission_amount=commission_amount,
                    salon_amount=salon_amount,
                    notes=str(notes).strip() if notes else None,
                )
            )
            counts["attendances"] += 1

    if "tbDespesas" in workbook.sheetnames:
        ws = workbook["tbDespesas"]
        for row in cell_values(ws):
            _, expense_date, category, description, amount, expense_type, paid_by = list(row[:7]) + [None] * max(0, 7 - len(row[:7]))
            if not description:
                continue
            db.session.add(
                Expense(
                    expense_date=parse_date(expense_date),
                    category=str(category or "Geral").strip(),
                    description=str(description).strip(),
                    amount=parse_float(amount),
                    expense_type=str(expense_type or "Variavel").strip(),
                    paid_by=str(paid_by).strip() if paid_by else None,
                )
            )
            counts["expenses"] += 1

    if "tbProdutos" in workbook.sheetnames:
        ws = workbook["tbProdutos"]
        for row in cell_values(ws):
            _, purchase_date, product_name, category, qty, unit_cost, total_cost, supplier = list(row[:8]) + [None] * max(0, 8 - len(row[:8]))
            if not product_name:
                continue
            quantity = parse_float(qty, 1)
            unit = parse_float(unit_cost)
            total = parse_float(total_cost, round(quantity * unit, 2))
            db.session.add(
                ProductPurchase(
                    purchase_date=parse_date(purchase_date),
                    product_name=str(product_name).strip(),
                    category=str(category or "Produto").strip(),
                    quantity=quantity,
                    unit_cost=unit,
                    total_cost=total,
                    supplier=str(supplier).strip() if supplier else None,
                )
            )
            counts["products"] += 1

    db.session.commit()
    return counts


def import_from_excel_file(file_storage, reset: bool = False) -> dict:
    workbook = openpyxl.load_workbook(file_storage, data_only=True)
    return import_workbook(workbook, reset=reset)


def import_from_excel_path(xlsx_path: str | Path, reset: bool = False) -> dict:
    workbook = openpyxl.load_workbook(Path(xlsx_path), data_only=True)
    return import_workbook(workbook, reset=reset)


MONTH_NAMES = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Março",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro",
}
MONTH_SHORT = {
    1: "Jan",
    2: "Fev",
    3: "Mar",
    4: "Abr",
    5: "Mai",
    6: "Jun",
    7: "Jul",
    8: "Ago",
    9: "Set",
    10: "Out",
    11: "Nov",
    12: "Dez",
}


@app.template_filter("currency_br")
def currency_br(value: float | None) -> str:
    amount = float(value or 0.0)
    formatted = f"{amount:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {formatted}"


@app.context_processor
def inject_globals() -> dict:
    today = date.today()
    return {
        "today": today,
        "current_year": today.year,
        "current_month": today.month,
        "month_names": MONTH_NAMES,
        "current_user": get_current_user(),
    }


@app.route("/login", methods=["GET", "POST"])
def login():
    if get_current_user():
        return redirect(url_for("index"))

    next_url = request.args.get("next") or request.form.get("next") or url_for("index")
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = User.query.filter_by(username=username, is_active=True).first()

        if user and user.check_password(password):
            session["user_id"] = user.id
            flash("Login realizado com sucesso.", "success")
            return redirect(next_url)

        flash("Usuário ou senha inválidos.", "error")

    return render_template("login.html", next_url=next_url, title="Login")


@app.route("/logout")
@login_required
def logout():
    session.pop("user_id", None)
    flash("Sessão encerrada.", "success")
    return redirect(url_for("login"))


@app.route("/")
@login_required
def index():
    year = parse_int(request.args.get("year"), date.today().year)
    month = parse_int(request.args.get("month"), date.today().month)
    summary = calculate_monthly_metrics(year, month)
    trend = calculate_recent_trend(summary["year"], summary["month"], 6)

    recent_attendances = Attendance.query.order_by(Attendance.attendance_date.desc(), Attendance.id.desc()).limit(8).all()
    recent_expenses = Expense.query.order_by(Expense.expense_date.desc(), Expense.id.desc()).limit(5).all()
    recent_products = ProductPurchase.query.order_by(ProductPurchase.purchase_date.desc(), ProductPurchase.id.desc()).limit(5).all()

    service_chart = summary["service_rank"][:6]
    collaborator_chart = summary["collaborator_rank"][:6]

    return render_template(
        "index.html",
        summary=summary,
        recent_attendances=recent_attendances,
        recent_expenses=recent_expenses,
        recent_products=recent_products,
        trend_labels=trend["labels"],
        trend_gross=trend["gross_values"],
        trend_result=trend["result_values"],
        service_labels=[row.service_name for row in service_chart],
        service_values=[float(row.gross_amount or 0) for row in service_chart],
        collaborator_labels=[row.collaborator_name for row in collaborator_chart],
        collaborator_values=[float(row.gross_amount or 0) for row in collaborator_chart],
    )


@app.route("/services", methods=["GET", "POST"])
@login_required
def services():
    editing_item = None
    edit_id = parse_int(request.args.get("edit"), 0)
    if edit_id:
        editing_item = Service.query.get_or_404(edit_id)

    if request.method == "POST":
        item_id = parse_int(request.form.get("item_id"), 0)
        if item_id:
            service = Service.query.get_or_404(item_id)
        else:
            service = Service()

        service.name = request.form.get("name", "").strip()
        service.category = request.form.get("category", "").strip() or "Geral"
        service.default_price = parse_float(request.form.get("default_price"))
        service.commission_pct = parse_float(request.form.get("commission_pct"))
        service.active = request.form.get("active") == "on"

        if not service.name:
            flash("Informe o nome do serviço.", "error")
            return redirect(url_for("services", edit=service.id) if item_id else url_for("services"))

        if not item_id:
            db.session.add(service)
        db.session.commit()
        flash("Serviço salvo com sucesso.", "success")
        return redirect(url_for("services"))

    items = Service.query.order_by(Service.active.desc(), Service.name.asc()).all()
    return render_template("services.html", items=items, editing_item=editing_item, title="Serviços")


@app.post("/services/<int:item_id>/delete")
@login_required
def service_delete(item_id: int):
    item = Service.query.get_or_404(item_id)
    if item.attendances:
        flash("Não dá para excluir um serviço que já possui atendimentos vinculados.", "error")
        return redirect(url_for("services"))
    db.session.delete(item)
    db.session.commit()
    flash("Serviço excluído.", "success")
    return redirect(url_for("services"))


@app.route("/collaborators", methods=["GET", "POST"])
@login_required
def collaborators():
    editing_item = None
    edit_id = parse_int(request.args.get("edit"), 0)
    if edit_id:
        editing_item = Collaborator.query.get_or_404(edit_id)

    if request.method == "POST":
        item_id = parse_int(request.form.get("item_id"), 0)
        if item_id:
            item = Collaborator.query.get_or_404(item_id)
        else:
            item = Collaborator()

        item.name = request.form.get("name", "").strip()
        item.role = request.form.get("role", "").strip() or "Profissional"
        item.default_commission_pct = parse_float(request.form.get("default_commission_pct"))
        item.status = request.form.get("status", "Ativo")

        if not item.name:
            flash("Informe o nome do colaborador.", "error")
            return redirect(url_for("collaborators", edit=item.id) if item_id else url_for("collaborators"))

        if not item_id:
            db.session.add(item)
        db.session.commit()
        flash("Colaborador salvo com sucesso.", "success")
        return redirect(url_for("collaborators"))

    items = Collaborator.query.order_by(Collaborator.status.asc(), Collaborator.name.asc()).all()
    return render_template("collaborators.html", items=items, editing_item=editing_item, title="Colaboradores")


@app.post("/collaborators/<int:item_id>/delete")
@login_required
def collaborator_delete(item_id: int):
    item = Collaborator.query.get_or_404(item_id)
    if item.attendances:
        flash("Não dá para excluir um colaborador que já possui atendimentos vinculados.", "error")
        return redirect(url_for("collaborators"))
    db.session.delete(item)
    db.session.commit()
    flash("Colaborador excluído.", "success")
    return redirect(url_for("collaborators"))


@app.route("/attendances", methods=["GET", "POST"])
@login_required
def attendances():
    services_list = Service.query.order_by(Service.active.desc(), Service.name.asc()).all()
    collaborators_list = Collaborator.query.order_by(Collaborator.status.asc(), Collaborator.name.asc()).all()

    editing_item = None
    edit_id = parse_int(request.args.get("edit"), 0)
    if edit_id:
        editing_item = Attendance.query.get_or_404(edit_id)

    if request.method == "POST":
        item_id = parse_int(request.form.get("item_id"), 0)
        payload = build_attendance_values(request.form)

        if item_id:
            item = Attendance.query.get_or_404(item_id)
        else:
            item = Attendance()

        for key, value in payload.items():
            setattr(item, key, value)

        if not item_id:
            db.session.add(item)
        db.session.commit()
        flash("Atendimento salvo com sucesso.", "success")
        return redirect(url_for("attendances"))

    items = Attendance.query.order_by(Attendance.attendance_date.desc(), Attendance.id.desc()).limit(100).all()
    return render_template(
        "attendances.html",
        items=items,
        services_list=services_list,
        collaborators_list=collaborators_list,
        editing_item=editing_item,
        title="Atendimentos",
    )


@app.post("/attendances/<int:item_id>/delete")
@login_required
def attendance_delete(item_id: int):
    item = Attendance.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    flash("Atendimento excluído.", "success")
    return redirect(url_for("attendances"))


@app.route("/expenses", methods=["GET", "POST"])
@login_required
def expenses():
    editing_item = None
    edit_id = parse_int(request.args.get("edit"), 0)
    if edit_id:
        editing_item = Expense.query.get_or_404(edit_id)

    if request.method == "POST":
        item_id = parse_int(request.form.get("item_id"), 0)
        if item_id:
            item = Expense.query.get_or_404(item_id)
        else:
            item = Expense()

        item.expense_date = parse_date(request.form.get("expense_date"))
        item.category = request.form.get("category", "").strip() or "Geral"
        item.description = request.form.get("description", "").strip()
        item.amount = parse_float(request.form.get("amount"))
        item.expense_type = request.form.get("expense_type", "Variavel").strip() or "Variavel"
        item.paid_by = request.form.get("paid_by", "").strip() or None

        if not item.description:
            flash("Informe a descrição da despesa.", "error")
            return redirect(url_for("expenses", edit=item.id) if item_id else url_for("expenses"))

        if not item_id:
            db.session.add(item)
        db.session.commit()
        flash("Despesa salva com sucesso.", "success")
        return redirect(url_for("expenses"))

    items = Expense.query.order_by(Expense.expense_date.desc(), Expense.id.desc()).limit(100).all()
    return render_template("expenses.html", items=items, editing_item=editing_item, title="Despesas")


@app.post("/expenses/<int:item_id>/delete")
@login_required
def expense_delete(item_id: int):
    item = Expense.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    flash("Despesa excluída.", "success")
    return redirect(url_for("expenses"))


@app.route("/products", methods=["GET", "POST"])
@login_required
def products():
    editing_item = None
    edit_id = parse_int(request.args.get("edit"), 0)
    if edit_id:
        editing_item = ProductPurchase.query.get_or_404(edit_id)

    if request.method == "POST":
        item_id = parse_int(request.form.get("item_id"), 0)
        if item_id:
            item = ProductPurchase.query.get_or_404(item_id)
        else:
            item = ProductPurchase()

        quantity = parse_float(request.form.get("quantity"), 1.0)
        unit_cost = parse_float(request.form.get("unit_cost"))
        total_cost_input = request.form.get("total_cost")
        total_cost = parse_float(total_cost_input, round(quantity * unit_cost, 2))
        if not total_cost_input or not str(total_cost_input).strip():
            total_cost = round(quantity * unit_cost, 2)

        item.purchase_date = parse_date(request.form.get("purchase_date"))
        item.product_name = request.form.get("product_name", "").strip()
        item.category = request.form.get("category", "").strip() or "Produto"
        item.quantity = quantity
        item.unit_cost = unit_cost
        item.total_cost = total_cost
        item.supplier = request.form.get("supplier", "").strip() or None

        if not item.product_name:
            flash("Informe o nome do produto.", "error")
            return redirect(url_for("products", edit=item.id) if item_id else url_for("products"))

        if not item_id:
            db.session.add(item)
        db.session.commit()
        flash("Compra de produto salva com sucesso.", "success")
        return redirect(url_for("products"))

    items = ProductPurchase.query.order_by(ProductPurchase.purchase_date.desc(), ProductPurchase.id.desc()).limit(100).all()
    return render_template("products.html", items=items, editing_item=editing_item, title="Produtos")


@app.post("/products/<int:item_id>/delete")
@login_required
def product_delete(item_id: int):
    item = ProductPurchase.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    flash("Compra de produto excluída.", "success")
    return redirect(url_for("products"))


@app.route("/report")
@login_required
def report():
    year = parse_int(request.args.get("year"), date.today().year)
    month = parse_int(request.args.get("month"), date.today().month)
    summary = calculate_monthly_metrics(year, month)
    return render_template("report.html", summary=summary, title="Relatório")


@app.route("/import-excel", methods=["GET", "POST"])
@login_required
def import_excel():
    if request.method == "POST":
        uploaded = request.files.get("xlsx_file")
        reset_data = request.form.get("reset_data") == "on"

        if not uploaded or not uploaded.filename:
            flash("Selecione um arquivo .xlsx para importar.", "error")
            return redirect(url_for("import_excel"))

        try:
            counts = import_from_excel_file(uploaded, reset=reset_data)
        except Exception as exc:
            flash(f"Falha ao importar planilha: {exc}", "error")
            return redirect(url_for("import_excel"))

        flash(
            "Importação concluída: "
            f"{counts['services']} serviços, "
            f"{counts['collaborators']} colaboradores, "
            f"{counts['attendances']} atendimentos, "
            f"{counts['expenses']} despesas e "
            f"{counts['products']} produtos.",
            "success",
        )
        return redirect(url_for("index"))

    return render_template("import_excel.html", title="Importar Excel")


@app.route("/export/monthly.xlsx")
@login_required
def export_monthly():
    year = parse_int(request.args.get("year"), date.today().year)
    month = parse_int(request.args.get("month"), date.today().month)
    summary = calculate_monthly_metrics(year, month)

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Resumo"
    ws_summary.append(["Métrica", "Valor"])
    ws_summary.append(["Ano", summary["year"]])
    ws_summary.append(["Mês", summary["month"]])
    ws_summary.append(["Entradas Brutas", summary["gross_revenue"]])
    ws_summary.append(["Comissões Pagas", summary["commissions_paid"]])
    ws_summary.append(["Receita Líquida Salão", summary["salon_revenue"]])
    ws_summary.append(["Despesas Operacionais", summary["operational_expenses"]])
    ws_summary.append(["Compras de Produtos", summary["product_costs"]])
    ws_summary.append(["Resultado do Mês", summary["result_month"]])
    ws_summary.append(["Qtd. Atendimentos", summary["count_attendances"]])
    ws_summary.append(["Ticket Médio", summary["average_ticket"]])

    ws_att = wb.create_sheet("Atendimentos")
    ws_att.append(
        [
            "Data",
            "Cliente",
            "Serviço",
            "Categoria",
            "Colaborador",
            "Forma de Pagamento",
            "Valor Cobrado",
            "% Comissão",
            "Comissão",
            "Valor Salão",
            "Observações",
        ]
    )
    for item in summary["attendances"]:
        ws_att.append(
            [
                item.attendance_date.strftime("%d/%m/%Y"),
                item.client_name or "",
                item.service.name,
                item.service.category,
                item.collaborator.name,
                item.payment_method,
                item.amount_charged,
                item.commission_pct_applied,
                item.commission_amount,
                item.salon_amount,
                item.notes or "",
            ]
        )

    ws_exp = wb.create_sheet("Despesas")
    ws_exp.append(["Data", "Categoria", "Descrição", "Valor", "Tipo", "Pago Por"])
    for item in summary["expenses"]:
        ws_exp.append(
            [
                item.expense_date.strftime("%d/%m/%Y"),
                item.category,
                item.description,
                item.amount,
                item.expense_type,
                item.paid_by or "",
            ]
        )

    ws_prod = wb.create_sheet("Produtos")
    ws_prod.append(["Data", "Produto", "Categoria", "Quantidade", "Custo Unitário", "Custo Total", "Fornecedor"])
    for item in summary["products"]:
        ws_prod.append(
            [
                item.purchase_date.strftime("%d/%m/%Y"),
                item.product_name,
                item.category,
                item.quantity,
                item.unit_cost,
                item.total_cost,
                item.supplier or "",
            ]
        )

    ws_service = wb.create_sheet("Ranking Servicos")
    ws_service.append(["Serviço", "Qtd. Atendimentos", "Faturamento Bruto", "Receita do Salão"])
    for row in summary["service_rank"]:
        ws_service.append([row.service_name, row.total_attendances, float(row.gross_amount or 0), float(row.salon_amount or 0)])

    ws_collab = wb.create_sheet("Ranking Colaboradores")
    ws_collab.append(["Colaborador", "Qtd. Atendimentos", "Faturamento Bruto", "Comissões", "Receita do Salão"])
    for row in summary["collaborator_rank"]:
        ws_collab.append(
            [
                row.collaborator_name,
                row.total_attendances,
                float(row.gross_amount or 0),
                float(row.commission_amount or 0),
                float(row.salon_amount or 0),
            ]
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"relatorio_mensal_{year}_{month:02d}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/seed-demo")
@login_required
def seed_demo():
    if Service.query.count() == 0:
        db.session.add_all(
            [
                Service(name="Corte Feminino", category="Cabelo", default_price=60, commission_pct=40, active=True),
                Service(name="Escova", category="Cabelo", default_price=45, commission_pct=35, active=True),
                Service(name="Pé e Mão", category="Unhas", default_price=55, commission_pct=50, active=True),
            ]
        )
    if Collaborator.query.count() == 0:
        db.session.add_all(
            [
                Collaborator(name="Ana", role="Cabeleireira", default_commission_pct=40, status="Ativo"),
                Collaborator(name="Bruna", role="Manicure", default_commission_pct=50, status="Ativo"),
            ]
        )
    db.session.commit()

    if Attendance.query.count() == 0:
        corte = Service.query.filter_by(name="Corte Feminino").first()
        ana = Collaborator.query.filter_by(name="Ana").first()
        pe_mao = Service.query.filter_by(name="Pé e Mão").first()
        bruna = Collaborator.query.filter_by(name="Bruna").first()
        if corte and ana:
            db.session.add(
                Attendance(
                    attendance_date=date.today(),
                    client_name="Cliente Demo",
                    service_id=corte.id,
                    collaborator_id=ana.id,
                    payment_method="Pix",
                    amount_charged=60,
                    commission_pct_applied=40,
                    commission_amount=24,
                    salon_amount=36,
                    notes="Dado de teste",
                )
            )
        if pe_mao and bruna:
            db.session.add(
                Attendance(
                    attendance_date=date.today(),
                    client_name="Cliente Demo 2",
                    service_id=pe_mao.id,
                    collaborator_id=bruna.id,
                    payment_method="Dinheiro",
                    amount_charged=55,
                    commission_pct_applied=50,
                    commission_amount=27.5,
                    salon_amount=27.5,
                    notes="Dado de teste",
                )
            )
        db.session.commit()

    if Expense.query.count() == 0:
        db.session.add(
            Expense(
                expense_date=date.today(),
                category="Operacional",
                description="Conta de energia",
                amount=180,
                expense_type="Fixa",
                paid_by="Caixa",
            )
        )
        db.session.commit()

    if ProductPurchase.query.count() == 0:
        db.session.add(
            ProductPurchase(
                purchase_date=date.today(),
                product_name="Shampoo Profissional",
                category="Estoque",
                quantity=3,
                unit_cost=25,
                total_cost=75,
                supplier="Fornecedor Demo",
            )
        )
        db.session.commit()

    flash("Dados de demonstração carregados.", "success")
    return redirect(url_for("index"))


with app.app_context():
    db.create_all()
    ensure_default_admin()


if __name__ == "__main__":
    app.run(debug=True)
